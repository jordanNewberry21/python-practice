# reading excel files

import openpyxl
import os

# change to the directory where the xcel file is located
os.chdir('/Users/jordannewberry/Desktop/')

# create a workbook object using openpyxl with the relative file path
workbook = openpyxl.load_workbook('example.xlsx')

# create a sheet object
sheet = workbook.get_sheet_by_name('Sheet1')

print(workbook.get_sheet_names())

print(sheet['A1'])

# create a cell object
cell = sheet['A1']
print(cell.value)

# without using cell variable
print(sheet['A1'].value)

print(sheet.cell(row=1, column=2).value)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
