# Editing Excel Spreadsheets

import openpyxl
wb = openpyxl.Workbook()

sheet = wb.get_sheet_by_name('Sheet')

print(sheet['A1'].value)

sheet['A2'] = 42
sheet['A1'] = 'Hello'
import os

os.chdir('/Users/jordannewberry/Desktop/')

wb.save('example.xlsx')

sheet2 = wb.create_sheet()

wb.get_sheet_names()

sheet2.title = 'My New Sheet Name'

print(wb.get_sheet_names())

wb.save('example2.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')

wb.save('example3.xlsx')
